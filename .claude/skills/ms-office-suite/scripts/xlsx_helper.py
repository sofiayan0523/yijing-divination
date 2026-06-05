#!/usr/bin/env python3
"""Excel spreadsheet helper — create, read, and edit .xlsx files.

Usage:
    python xlsx_helper.py create <output.xlsx> <sheet_title>
    python xlsx_helper.py read <input.xlsx> [sheet_name]
    python xlsx_helper.py info <input.xlsx>
"""
import sys
import json


def create_workbook(output_path: str, sheet_title: str = "Sheet1") -> None:
    """Create a simple workbook with headers as a starting point."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    wb = Workbook()
    ws = wb.active
    ws.title = sheet_title

    # Example header row
    header_font = Font(bold=True, size=11, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="2F5496")
    header_align = Alignment(horizontal="center")

    headers = ["Column A", "Column B", "Column C"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align

    ws.freeze_panes = "A2"
    ws.column_dimensions["A"].width = 15
    ws.column_dimensions["B"].width = 15
    ws.column_dimensions["C"].width = 15

    wb.save(output_path)
    print(f"Created: {output_path}")


def read_workbook(input_path: str, sheet_name: str | None = None) -> None:
    """Read and print spreadsheet contents."""
    from openpyxl import load_workbook

    wb = load_workbook(input_path, data_only=True)
    sheets = [sheet_name] if sheet_name else wb.sheetnames

    for name in sheets:
        if name not in wb.sheetnames:
            print(f"Sheet '{name}' not found. Available: {wb.sheetnames}")
            continue
        ws = wb[name]
        print(f"\n=== {name} ({ws.max_row} rows x {ws.max_column} cols) ===")
        for row in ws.iter_rows(values_only=True):
            values = [str(v) if v is not None else "" for v in row]
            print(" | ".join(values))


def get_info(input_path: str) -> None:
    """Get workbook metadata as JSON."""
    from openpyxl import load_workbook

    wb = load_workbook(input_path, read_only=True, data_only=True)
    info = {
        "sheet_count": len(wb.sheetnames),
        "sheets": [],
    }
    for name in wb.sheetnames:
        ws = wb[name]
        info["sheets"].append({
            "name": name,
            "max_row": ws.max_row,
            "max_column": ws.max_column,
        })
    wb.close()
    print(json.dumps(info, indent=2))


if __name__ == "__main__":
    if len(sys.argv) < 3:
        print(__doc__)
        sys.exit(1)

    cmd = sys.argv[1]
    if cmd == "create":
        create_workbook(
            sys.argv[2],
            sys.argv[3] if len(sys.argv) > 3 else "Sheet1",
        )
    elif cmd == "read":
        read_workbook(
            sys.argv[2],
            sys.argv[3] if len(sys.argv) > 3 else None,
        )
    elif cmd == "info":
        get_info(sys.argv[2])
    else:
        print(f"Unknown command: {cmd}")
        sys.exit(1)
